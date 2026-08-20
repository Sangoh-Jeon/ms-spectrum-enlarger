"""
desktop_app.py — 데스크톱 GUI (Tkinter) 전용 실행 파일
====================================================
웹 앱(Streamlit)과 독립적으로 로컬 PC에서 실행할 때 사용합니다.

실행 방법:
    python desktop_app.py
    python desktop_app.py your_file.xlsx   # CLI 모드
"""
import os
import sys
import concurrent.futures
import openpyxl

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except ImportError:
    print("Tkinter를 사용할 수 없는 환경입니다 (서버/클라우드 환경).")
    sys.exit(1)

from enlarge_spectrum_peaks import (
    get_base_dir,
    extract_peaks_with_google_vision,
    process_excel_with_selections,
)


class MRMSelectionModal(tk.Toplevel):
    """
    Modal Dialog popup showing extracted m/z peak values per Sheet and Graph Type (Precursor vs Product Ion).
    Features:
    - Smart default auto-selection (Top 1 Precursor, Top 3 Product).
    - Editable m/z text boxes so users can manually correct any misread OCR numbers!
    """

    def __init__(self, parent, excel_path, sheet_peak_data, on_confirm_callback):
        super().__init__(parent)
        self.title("MRM 조건 이온 선택 및 수기 수정")
        self.geometry("820x660")
        self.transient(parent)

        self.excel_path = excel_path
        self.sheet_peak_data = sheet_peak_data
        self.on_confirm_callback = on_confirm_callback
        self.peak_widgets = {}  # {sheet_name: {'precursor': [(var, entry, orig_mz)], 'product': ...}}

        self._build_ui()
        self.grab_set()
        self.focus_set()

    def _build_ui(self):
        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(
            main_frame,
            text="MRM 조건 이온 선택 및 분자량(m/z) 수기 수정",
            font=("맑은 고딕", 13, "bold"),
            foreground="navy"
        ).pack(anchor=tk.W, pady=(0, 5))

        guide_text = (
            "✓ OCR이 오인식한 분자량이 있다면 입력창에서 직접 수기로 수정할 수 있습니다.\n"
            "✓ 스마트 자동 체크: Precursor 상위 1개, Product 상위 3개 피크 기본 선택.\n"
            "✓ 체크한 이온: 파란색 (Blue) 48pt 강조 / 체크 안 한 이온: 회색 (Gray) 46pt 표시."
        )
        ttk.Label(main_frame, text=guide_text, font=("맑은 고딕", 10), foreground="#333333").pack(anchor=tk.W, pady=(0, 10))

        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=5)

        for sheet_name, graph_data in self.sheet_peak_data.items():
            sheet_tab = ttk.Frame(notebook, padding="10")
            notebook.add(sheet_tab, text=f" 📄 {sheet_name} ")

            self.peak_widgets[sheet_name] = {'precursor': [], 'product': []}

            canvas = tk.Canvas(sheet_tab, borderwidth=0, highlightthickness=0)
            scrollbar = ttk.Scrollbar(sheet_tab, orient="vertical", command=canvas.yview)
            scroll_content = ttk.Frame(canvas, padding="5")

            canvas_window = canvas.create_window((0, 0), window=scroll_content, anchor="nw")

            def _on_canvas_resize(event, c=canvas, w=canvas_window):
                c.itemconfig(w, width=event.width)

            def _on_content_resize(event, c=canvas):
                c.configure(scrollregion=c.bbox('all'))

            canvas.bind('<Configure>', _on_canvas_resize)
            scroll_content.bind('<Configure>', _on_content_resize)
            canvas.configure(yscrollcommand=scrollbar.set)
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            # Precursor Ion Frame
            self._build_ion_frame(
                scroll_content, sheet_name,
                ion_type='precursor',
                graph_data=graph_data,
                frame_label=" 🔹 Precursor Ion 피크 (입력창에서 수기 수정 가능) ",
                star_tag="Da ⭐ (최고 피크)"
            )

            # Product Ion Frame
            self._build_ion_frame(
                scroll_content, sheet_name,
                ion_type='product',
                graph_data=graph_data,
                frame_label=" 🔸 Product Ion 피크 (입력창에서 수기 수정 가능) ",
                star_tag="Da ⭐ (추천)"
            )

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))

        ttk.Button(btn_frame, text="✓ MRM 이온 선택 완료 및 엑셀 생성", command=self._on_confirm).pack(side=tk.RIGHT, ipadx=10, ipady=5)
        ttk.Button(btn_frame, text="취소", command=self.destroy).pack(side=tk.RIGHT, padx=10, ipady=5)

        self.update_idletasks()

    def _build_ion_frame(self, parent, sheet_name, ion_type, graph_data, frame_label, star_tag):
        """Precursor / Product 공통 이온 프레임 빌더 (중복 코드 통합)."""
        lf = ttk.LabelFrame(parent, text=frame_label, padding="10")
        lf.pack(fill=tk.X, expand=True, pady=5 if ion_type == 'precursor' else 10, padx=5)

        info = graph_data.get(ion_type, {'all_peaks': [], 'default_checked': set()})
        peaks = info['all_peaks']
        defaults = info['default_checked']

        if not peaks:
            ttk.Label(lf, text="검출된 피크가 없습니다.", font=("맑은 고딕", 9, "italic")).pack(anchor=tk.W)
            return

        grid_frame = ttk.Frame(lf)
        grid_frame.pack(fill=tk.X, expand=True)

        for idx, mz in enumerate(peaks):
            is_checked = mz in defaults
            var = tk.BooleanVar(value=is_checked)

            row = idx // 2
            col_base = (idx % 2) * 3

            ttk.Checkbutton(grid_frame, variable=var).grid(row=row, column=col_base, sticky=tk.W, padx=(10, 2), pady=4)

            entry = ttk.Entry(grid_frame, width=10, font=("맑은 고딕", 10, "bold"))
            entry.insert(0, mz)
            entry.grid(row=row, column=col_base + 1, sticky=tk.W, padx=(0, 2), pady=4)

            tag = star_tag if is_checked else "Da"
            ttk.Label(grid_frame, text=tag, font=("맑은 고딕", 9)).grid(row=row, column=col_base + 2, sticky=tk.W, padx=(0, 25), pady=4)

            self.peak_widgets[sheet_name][ion_type].append((var, entry, mz))

    def _on_confirm(self):
        selections = {}
        for sheet_name, cat_dict in self.peak_widgets.items():
            selections[sheet_name] = {'precursor': [], 'product': []}
            for category in ['precursor', 'product']:
                for var, entry, orig_mz in cat_dict[category]:
                    user_val = entry.get().strip()
                    try:
                        formatted_val = f"{float(user_val):.2f}"
                    except ValueError:
                        formatted_val = user_val

                    selections[sheet_name][category].append({
                        'orig_mz': orig_mz,
                        'final_mz': formatted_val,
                        'is_mrm': var.get()
                    })
        self.destroy()
        self.on_confirm_callback(self.excel_path, selections)


class AppGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("GLP 보고서용 질량분석 피크 분자량(m/z) 확대 프로그램")
        self.root.geometry("680x420")
        self.root.resizable(False, False)

        style = ttk.Style()
        style.theme_use('clam')

        frame = ttk.Frame(root, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="질량분석(MS/MS) 피크 분자량(m/z) 글자 확대기", font=("맑은 고딕", 14, "bold")).pack(anchor=tk.W, pady=(0, 5))
        ttk.Label(frame, text="Data 폴더 선택 후 팝업창에서 MRM 이온 체크 및 수기 수정이 가능합니다.", font=("맑은 고딕", 9)).pack(anchor=tk.W, pady=(0, 15))

        # 1. 파일 선택
        file_box = ttk.LabelFrame(frame, text=" 📂 대상 엑셀 파일 선택 (Data 폴더) ", padding="12")
        file_box.pack(fill=tk.X, pady=5)

        self.data_dir = os.path.join(get_base_dir(), "Data")
        os.makedirs(self.data_dir, exist_ok=True)

        self.combo_files = ttk.Combobox(file_box, state="readonly", font=("맑은 고딕", 10))
        self.combo_files.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        ttk.Button(file_box, text="🔄 목록 새로고침", command=self.refresh_file_list).pack(side=tk.RIGHT)
        self.refresh_file_list()

        # 2. 옵션
        opt_frame = ttk.LabelFrame(frame, text=" 확대 폰트 설정 ", padding="10")
        opt_frame.pack(fill=tk.X, pady=10)

        ttk.Label(opt_frame, text="MRM 강조 폰트 크기:").pack(side=tk.LEFT, padx=(0, 10))
        self.font_size_var = tk.IntVar(value=48)
        ttk.Spinbox(opt_frame, from_=20, to=72, increment=4, textvariable=self.font_size_var, width=5).pack(side=tk.LEFT)
        ttk.Label(opt_frame, text="pt (MRM 선택 이온: 파란색 48pt / 기타 이온: 회색 46pt)").pack(side=tk.LEFT, padx=(5, 0))

        # 3. 상태 표시
        self.status_var = tk.StringVar(value="Data 폴더의 엑셀 파일을 선택한 후 '분자량 피크 읽기 및 확대' 버튼을 클릭하세요.")
        ttk.Label(frame, textvariable=self.status_var, font=("맑은 고딕", 9), foreground="blue").pack(anchor=tk.W, pady=10)

        # 4. 실행 버튼
        self.btn_run = ttk.Button(frame, text="🔍 1단계: 분자량 피크 읽기 & MRM 이온 선택/수정 팝업 열기", command=self.start_mrm_selection)
        self.btn_run.pack(fill=tk.X, ipady=8, pady=(5, 0))

    def refresh_file_list(self):
        """Data 폴더와 루트 폴더에서 엑셀 파일 목록을 갱신합니다."""
        exclude_suffixes = ('_확대.xlsx', '_분자량확대.xlsx')
        file_list = []
        base_d = get_base_dir()

        for folder in [self.data_dir, base_d]:
            if os.path.exists(folder):
                for f in os.listdir(folder):
                    if f.endswith(('.xlsx', '.xls')) and not any(f.endswith(s) for s in exclude_suffixes) and f not in file_list:
                        file_list.append(f)

        self.combo_files['values'] = file_list
        if file_list:
            self.combo_files.current(0)
        else:
            self.combo_files.set('')

    def update_status(self, msg):
        self.status_var.set(msg)
        self.root.update()

    def start_mrm_selection(self):
        selected_filename = self.combo_files.get().strip()
        if not selected_filename:
            messagebox.showwarning("경고", "처리할 엑셀 파일을 선택하세요.")
            return

        excel_path = os.path.join(self.data_dir, selected_filename)
        if not os.path.exists(excel_path):
            excel_path = os.path.join(get_base_dir(), selected_filename)

        if not os.path.exists(excel_path):
            messagebox.showerror("오류", f"파일을 찾을 수 없습니다: {excel_path}")
            return

        try:
            self.btn_run.config(state=tk.DISABLED)
            self.update_status("🔍 엑셀 내 Precursor / Product Ion 분자량 피크 읽는 중...")

            wb = openpyxl.load_workbook(excel_path)
            sheet_peak_data = {
                name: {
                    'precursor': {'all_peaks': [], 'default_checked': set()},
                    'product': {'all_peaks': [], 'default_checked': set()}
                }
                for name in wb.sheetnames
            }

            tasks = [
                (name, idx % 2 == 0, img._data())
                for name in wb.sheetnames
                for idx, img in enumerate(wb[name]._images)
            ]
            total_images = len(tasks)
            proc_img_count = 0

            def _proc_task(task_item):
                s_name, is_prec, data_bytes = task_item
                res = extract_peaks_with_google_vision(data_bytes, is_precursor=is_prec)
                return s_name, is_prec, res

            max_workers = min(4, os.cpu_count() or 4)
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_task = {executor.submit(_proc_task, t): t for t in tasks}
                for future in concurrent.futures.as_completed(future_to_task):
                    proc_img_count += 1
                    s_name, is_prec, peak_info = future.result()
                    self.update_status(f"🔍 고속 피크 분석 중 ({proc_img_count}/{total_images}) - 시트: {s_name}")

                    cat_key = 'precursor' if is_prec else 'product'
                    sheet_peak_data[s_name][cat_key]['all_peaks'].extend(peak_info['all_peaks'])
                    sheet_peak_data[s_name][cat_key]['default_checked'].update(peak_info['default_checked'])

            for name in wb.sheetnames:
                for cat in ('precursor', 'product'):
                    sheet_peak_data[name][cat]['all_peaks'] = sorted(
                        list(set(sheet_peak_data[name][cat]['all_peaks'])),
                        key=lambda x: float(x)
                    )

            self.update_status("✅ 초고속 피크 추출 완료! MRM 이온 선택/수정 팝업창을 확인하세요.")
            MRMSelectionModal(self.root, excel_path, sheet_peak_data, self.execute_excel_generation)

        except Exception as e:
            self.update_status("❌ 이미지 피크 읽기 오류 발생")
            messagebox.showerror("오류", f"피크 읽기 중 오류가 발생했습니다:\n{str(e)}")
        finally:
            self.btn_run.config(state=tk.NORMAL)

    def execute_excel_generation(self, excel_path, sheet_selections):
        try:
            self.btn_run.config(state=tk.DISABLED)
            font_size = int(self.font_size_var.get())
            out_file = process_excel_with_selections(excel_path, sheet_selections, font_size=font_size, status_callback=self.update_status)

            self.status_var.set(f"🎉 완벽 완료! 저장됨: {os.path.basename(out_file)}")
            messagebox.showinfo("성공", f"수정된 MRM 이온(파란색 강조)이 적용된 최종 엑셀 파일이 저장되었습니다!\n\n저장 경로:\n{out_file}")
        except Exception as e:
            self.status_var.set("❌ 엑셀 저장 중 오류 발생")
            messagebox.showerror("오류", f"엑셀 생성 중 오류가 발생했습니다:\n{str(e)}")
        finally:
            self.btn_run.config(state=tk.NORMAL)


def main():
    if len(sys.argv) > 1 and sys.argv[1].endswith(('.xlsx', '.xls')):
        excel_file = sys.argv[1]
        print(f"CLI 모드로 실행합니다: {excel_file}")
        wb = openpyxl.load_workbook(excel_file)
        sheet_selections = {name: {'precursor': [], 'product': []} for name in wb.sheetnames}
        out = process_excel_with_selections(excel_file, sheet_selections, font_size=48, status_callback=print)
        print(f"완료! 저장 위치: {out}")
    else:
        root = tk.Tk()
        AppGUI(root)
        root.mainloop()


if __name__ == "__main__":
    main()
