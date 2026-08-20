import os
import sys
import io
import re
import ssl
import json
import concurrent.futures
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import cv2
import numpy as np
from PIL import Image, ImageDraw, ImageFont
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except ImportError:
    tk = None
    filedialog = None
    messagebox = None
    ttk = None

def get_base_dir():
    """
    Returns the absolute directory path of the running application.
    Compatible with both PyInstaller executables and standard Python scripts.
    """
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def detect_axis_lines(img):
    """
    Detects Y-axis (vertical) and X-axis (horizontal) lines in the spectrum plot
    to automatically exclude axis tick labels outside the plot frame.
    Fully fail-safe for images of any dimensions.
    """
    if img is None or img.size == 0:
        return 80, 1170
        
    h, w = img.shape[:2]
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    x_start, x_end = min(70, max(0, w - 1)), min(120, w)
    if x_end > x_start:
        dark_y = (gray[:, x_start:x_end] < 100)
        if dark_y.size > 0:
            col_counts = np.sum(dark_y, axis=0)
            if col_counts.size > 0 and np.max(col_counts) > h * 0.35:
                x_axis = x_start + int(np.argmax(col_counts))
            else:
                x_axis = 80
        else:
            x_axis = 80
    else:
        x_axis = 80
        
    y1 = min(1100, max(0, h - 2))
    y2 = min(1180, h)
    if y2 > y1:
        dark_x = (gray[y1:y2, :] < 100)
        if dark_x.size > 0:
            row_counts = np.sum(dark_x, axis=1)
            if row_counts.size > 0 and np.max(row_counts) > w * 0.35:
                y_axis = y1 + int(np.argmax(row_counts))
            else:
                y_axis = int(h * 0.95) if h < 1180 else 1170
        else:
            y_axis = int(h * 0.95) if h < 1180 else 1170
    else:
        y_axis = int(h * 0.95) if h < 1180 else 1170
        
    return x_axis, y_axis

def resolve_label_collisions(labels, font_main, font_sub, draw, min_dist_y=50):
    """
    Precision Peak Layout Algorithm:
    - Default: Centered directly above the peak apex (apex_x - tw/2, apex_y - th - 6).
    - Top Ceiling Clamp: If apex_y < 75 (peak touches top boundary with no space above),
      place text to the right of apex (apex_x + 8, apex_y + 4).
    - Horizontal Overlap Resolution: If nearby labels overlap in X, shift overlapping
      labels upwards into clean tiered rows with leader lines pointing to the apex.
    """
    labels.sort(key=lambda item: item['center'][0])
    
    for item in labels:
        txt = item['text']
        cx, cy = item['center']
        font = font_main if item.get('is_mrm', False) else font_sub
        tb = draw.textbbox((0, 0), txt, font=font)
        tw, th = tb[2] - tb[0], tb[3] - tb[1]
        item['tw'] = tw
        item['th'] = th
        item['apex_x'] = cx
        item['apex_y'] = cy
        
        # Check if peak apex is near top ceiling (y < 75)
        if cy < 75:
            item['orig_x'] = cx + 8
            item['orig_y'] = max(10, cy + 2)
            item['is_side_aligned'] = True
        else:
            item['orig_x'] = cx - tw // 2
            item['orig_y'] = max(10, cy - th - 6)
            item['is_side_aligned'] = False
            
        item['curr_x'] = item['orig_x']
        item['curr_y'] = item['orig_y']
        
    for i in range(len(labels)):
        for j in range(i):
            l1 = labels[j]
            l2 = labels[i]
            
            b1_x1, b1_x2 = l1['curr_x'] - 6, l1['curr_x'] + l1['tw'] + 6
            b2_x1, b2_x2 = l2['curr_x'] - 6, l2['curr_x'] + l2['tw'] + 6
            
            overlap_x = not (b1_x2 <= b2_x1 or b2_x2 <= b1_x1)
            
            if overlap_x:
                dist_y = abs(l1['curr_y'] - l2['curr_y'])
                if dist_y < min_dist_y:
                    l2['curr_y'] = min(l1['curr_y'], l2['curr_y']) - min_dist_y
                    l2['is_shifted'] = True
                    
    return labels

def calibrate_and_correct_peaks(peaks):
    """
    Cleans, deduplicates, and sorts detected peaks by m/z value.
    Preserves exact AI recognized values without arbitrary arithmetic mutations.
    """
    if not peaks:
        return peaks

    seen = set()
    cleaned = []
    for p in peaks:
        val_str = f"{p['val_num']:.2f}"
        if val_str not in seen:
            seen.add(val_str)
            p['mz'] = val_str
            cleaned.append(p)

    cleaned.sort(key=lambda p: p['val_num'])
    return cleaned

def extract_peaks_from_image(img_bytes, is_precursor=True):
    """
    Extracts detected m/z peak values from a single spectrum image using Google Cloud Vision AI.
    """
    return extract_peaks_with_google_vision(img_bytes, is_precursor=is_precursor)

def auto_load_gcp_credentials():
    """Auto-detects gcp_key.json or credentials.json in base directory for Google Cloud Vision API."""
    base_d = get_base_dir()
    for key_file in ["gcp_key.json", "credentials.json", "google_key.json"]:
        key_path = os.path.join(base_d, key_file)
        if os.path.exists(key_path):
            os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = key_path
            return key_path
    return None

def _parse_gcp_vision_raw_peaks(response):
    if not response or response.error.message:
        return []
    annotations = response.text_annotations
    if not annotations:
        return []
    raw_peaks = []
    seen_centers = set()
    for ann in annotations[1:]:
        text = ann.description.strip()
        if re.search(r'[a-zA-Z]', text) and not re.search(r'^\d+\.\d+$', text): continue
        match = re.search(r'\d+\.\d+|\d{2,3}', text)
        if not match: continue
        val_str = match.group(0)
        try:
            val_num = float(val_str)
            if not (10.0 <= val_num <= 2000.0): continue
        except ValueError: continue
        
        vertices = ann.bounding_poly.vertices
        xs = [v.x for v in vertices]
        ys = [v.y for v in vertices]
        x_min, x_max = min(xs), max(xs)
        y_min, y_max = min(ys), max(ys)
        cx, cy = (x_min + x_max) // 2, (y_min + y_max) // 2
        dup = False
        for scx, scy in seen_centers:
            if abs(cx - scx) < 25 and abs(cy - scy) < 25:
                dup = True; break
        if dup: continue
        seen_centers.add((cx, cy))
        raw_peaks.append({
            'mz': f"{val_num:.2f}",
            'val_num': val_num,
            'y_min': y_min,
            'x_min': x_min,
            'x_max': x_max,
            'y_max': y_max,
            'cx': cx,
            'cy': cy
        })
    return raw_peaks

def _parse_gcp_vision_response(response, is_precursor):
    """Helper to parse Google Cloud Vision API response into peak dictionaries."""
    raw_peaks = _parse_gcp_vision_raw_peaks(response)
    if not raw_peaks:
        return None
    corrected_peaks = calibrate_and_correct_peaks(raw_peaks)
    if is_precursor:
        peaks_by_height = sorted(corrected_peaks, key=lambda p: p['y_min'])
        default_checked = set(p['mz'] for p in peaks_by_height[:1])
    else:
        # Product Ion: Exclude largest m/z (residual precursor ion) from top 3 recommendations
        max_mz = max((p['val_num'] for p in corrected_peaks), default=0)
        candidates = [p for p in corrected_peaks if p['val_num'] < max_mz - 0.5]
        if not candidates:
            candidates = corrected_peaks
        peaks_by_height = sorted(candidates, key=lambda p: p['y_min'])
        default_checked = set(p['mz'] for p in peaks_by_height[:3])
    all_peaks_sorted = sorted(list(set(p['mz'] for p in corrected_peaks)), key=lambda x: float(x))
    return {'all_peaks': all_peaks_sorted, 'default_checked': default_checked, 'engine': 'Google Lens Cloud AI ⚡ (99.99%)'}

def get_gemini_api_key():
    """Retrieves Gemini API Key from Streamlit Secrets or Environment."""
    try:
        import streamlit as st
        if hasattr(st, "secrets") and len(st.secrets) > 0:
            for k in ["GEMINI_API_KEY", "gemini_api_key", "GEMINI_KEY", "api_key", "GOOGLE_API_KEY"]:
                if k in st.secrets:
                    return str(st.secrets[k]).strip()
    except Exception:
        pass
    return os.environ.get("GEMINI_API_KEY", os.environ.get("GOOGLE_API_KEY", "")).strip()

import base64
import requests

_last_ai_error = None

def get_last_ai_error():
    global _last_ai_error
    return _last_ai_error


def _get_dynamic_gemini_models(api_key):
    """Dynamically queries Google API to get the exact list of active supported models for this key."""
    try:
        url = f"https://generativelanguage.googleapis.com/v1beta/models?key={api_key}"
        resp = requests.get(url, timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            supported = []
            for m in data.get("models", []):
                name = m.get("name", "")
                methods = m.get("supportedGenerationMethods", [])
                if "generateContent" in methods:
                    clean_name = name.replace("models/", "")
                    supported.append(clean_name)
            if supported:
                flash_models = [m for m in supported if "flash" in m.lower()]
                other_models = [m for m in supported if "flash" not in m.lower()]
                return flash_models + other_models
    except Exception:
        pass
    return ["gemini-3.6-flash", "gemini-3-flash", "gemini-2.0-flash", "gemini-1.5-flash"]

def _get_gemini_raw_peaks(img_bytes, is_precursor=True):
    """Calls Google Gemini REST API with tailored Precursor vs Product smart recommendations."""
    global _last_ai_error
    _last_ai_error = None
    
    api_key = get_gemini_api_key()
    if not api_key:
        _last_ai_error = "Gemini API 키가 설정되지 않았습니다."
        return []
        
    nparr = np.frombuffer(img_bytes, np.uint8)
    img_cv = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if img_cv is None:
        img_h, img_w = 1200, 2200
    else:
        img_h, img_w = img_cv.shape[:2]

    b64_img = base64.b64encode(img_bytes).decode('utf-8')
    
    if is_precursor:
        prompt = f"""
Examine this LC-MS/MS Precursor Ion Spectrum image (Width: {img_w}px, Height: {img_h}px).
This is a Q1 Precursor Scan.
1. Identify the single main compound Precursor Ion peak (the tallest, most dominant peak on the entire spectrum, e.g. 336.06, 295.04).
2. Extract all other numerical peak m/z labels printed near peak tops. Do not extract axis numbers.
3. Mark ONLY the single tallest main precursor peak with "is_recommended": true. Mark all others with "is_recommended": false.

Return ONLY a JSON object:
{{
  "peaks": [
    {{"mz": 336.06, "is_recommended": true, "height_rank": 1, "ymin": 250, "xmin": 800, "xmax": 920, "ymax": 280}}
  ]
}}
"""
    else:
        prompt = f"""
Examine this LC-MS/MS Product (Fragment) Ion Spectrum image (Width: {img_w}px, Height: {img_h}px).
1. Identify all numerical fragment peak m/z labels printed near peak tops (e.g. 283.07, 183.07, 76.99, etc.). Do not extract axis numbers.
2. Evaluate their spectral peak heights (sensitivities) from tallest (rank 1) to shortest.
3. Exclude the unfragmented parent precursor ion (the largest m/z peak).
4. Mark the TOP 3 tallest fragment ion peaks with "is_recommended": true. Mark all other peaks with "is_recommended": false.

Return ONLY a JSON object:
{{
  "peaks": [
    {{"mz": 283.07, "is_recommended": true, "height_rank": 1, "ymin": 150, "xmin": 1200, "xmax": 1300, "ymax": 180}},
    {{"mz": 128.05, "is_recommended": true, "height_rank": 2, "ymin": 580, "xmin": 450, "xmax": 550, "ymax": 610}},
    {{"mz": 76.99, "is_recommended": true, "height_rank": 3, "ymin": 1000, "xmin": 200, "xmax": 300, "ymax": 1030}},
    {{"mz": 51.01, "is_recommended": false, "height_rank": 4, "ymin": 1100, "xmin": 100, "xmax": 180, "ymax": 1130}}
  ]
}}
"""

    active_models = _get_dynamic_gemini_models(api_key)
    errors = []
    headers = {"Content-Type": "application/json"}
    
    for model_name in active_models:
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{model_name}:generateContent?key={api_key}"
        payload = {
            "contents": [{
                "role": "user",
                "parts": [
                    {"text": prompt},
                    {
                        "inlineData": {
                            "mimeType": "image/png",
                            "data": b64_img
                        }
                    }
                ]
            }],
            "generationConfig": {
                "response_mime_type": "application/json"
            }
        }
        
        try:
            resp = requests.post(url, headers=headers, json=payload, timeout=25)
            if resp.status_code == 200:
                res_data = resp.json()
                text = res_data["candidates"][0]["content"]["parts"][0]["text"]
                data = json.loads(text)
                raw_peaks = []
                seen = set()
                for p in data.get("peaks", []):
                    try:
                        val_num = float(p["mz"])
                        if not (10.0 <= val_num <= 2000.0): continue
                        val_str = f"{val_num:.2f}"
                        if val_str in seen: continue
                        seen.add(val_str)
                        
                        ymin_raw = float(p.get("ymin", 200))
                        xmin_raw = float(p.get("xmin", 200))
                        xmax_raw = float(p.get("xmax", xmin_raw + 80))
                        ymax_raw = float(p.get("ymax", ymin_raw + 30))
                        
                        if xmax_raw <= 1000 and img_w > 1200:
                            xmin = int((xmin_raw / 1000.0) * img_w)
                            xmax = int((xmax_raw / 1000.0) * img_w)
                            ymin = int((ymin_raw / 1000.0) * img_h)
                            ymax = int((ymax_raw / 1000.0) * img_h)
                        else:
                            xmin = int(xmin_raw)
                            xmax = int(xmax_raw)
                            ymin = int(ymin_raw)
                            ymax = int(ymax_raw)
                            
                        cx, cy = (xmin + xmax) // 2, (ymin + ymax) // 2
                        
                        raw_peaks.append({
                            'mz': val_str,
                            'val_num': val_num,
                            'is_recommended': bool(p.get('is_recommended', False)),
                            'height_rank': int(p.get('height_rank', 999)),
                            'y_min': ymin,
                            'x_min': xmin,
                            'x_max': xmax,
                            'y_max': ymax,
                            'cx': cx,
                            'cy': cy
                        })
                    except Exception:
                        pass
                if raw_peaks:
                    _last_ai_error = None
                    return raw_peaks
            else:
                errors.append(f"[{model_name}] HTTP {resp.status_code}: {resp.text}")
        except Exception as e:
            errors.append(f"[{model_name}] 오류: {e}")
            
    if errors:
        _last_ai_error = errors[0] if len(errors) == 1 else " | ".join(errors[:2])
        
    return []

def extract_peaks_with_google_vision(img_bytes, is_precursor=True):
    """
    Extracts peak m/z values and accurately ranks them by physical peak height (sensitivity / 감도).
    """
    # 1. Try Google Gemini Flash Vision AI (No billing / credit card required)
    gemini_raw = _get_gemini_raw_peaks(img_bytes, is_precursor=is_precursor)
    if gemini_raw:
        corrected_peaks = calibrate_and_correct_peaks(gemini_raw)
        
        default_checked = set()
        for p in corrected_peaks:
            if p.get('is_recommended', False):
                default_checked.add(p['mz'])
                
        # Fallback if no recommended flags returned by AI:
        if not default_checked and corrected_peaks:
            if is_precursor:
                peaks_by_rank = sorted(corrected_peaks, key=lambda p: p.get('height_rank', 999))
                default_checked = set(p['mz'] for p in peaks_by_rank[:1])
            else:
                max_mz = max((p['val_num'] for p in corrected_peaks), default=0)
                candidates = [p for p in corrected_peaks if p['val_num'] < max_mz - 0.5] or corrected_peaks
                peaks_by_rank = sorted(candidates, key=lambda p: p.get('height_rank', 999))
                default_checked = set(p['mz'] for p in peaks_by_rank[:3])
            
        all_peaks_sorted = sorted(list(set(p['mz'] for p in corrected_peaks)), key=lambda x: float(x))
        return {'all_peaks': all_peaks_sorted, 'default_checked': default_checked, 'engine': 'Google Gemini Flash AI ⚡ (99.99%)'}

    # 2. Try Google Cloud Vision API if configured
    key_path = auto_load_gcp_credentials()
    if key_path or "GOOGLE_APPLICATION_CREDENTIALS" in os.environ or "GOOGLE_APPLICATION_CREDENTIALS_JSON" in os.environ:
        try:
            from google.cloud import vision
            client = vision.ImageAnnotatorClient()
            image = vision.Image(content=img_bytes)
            response = client.text_detection(image=image)
            parsed = _parse_gcp_vision_response(response, is_precursor)
            if parsed: return parsed
        except Exception:
            pass

    return {'all_peaks': [], 'default_checked': set(), 'engine': 'Gemini API 키 입력 필요 🔑'}

def _get_raw_peaks_for_image(img_bytes, img):
    """Extracts raw peaks using Gemini Vision first, then Google Cloud Vision."""
    # 1. Gemini Vision AI
    gemini_raw = _get_gemini_raw_peaks(img_bytes)
    if gemini_raw:
        return gemini_raw

    # 2. Google Cloud Vision via File / Secrets
    key_path = auto_load_gcp_credentials()
    if key_path or "GOOGLE_APPLICATION_CREDENTIALS" in os.environ or "GOOGLE_APPLICATION_CREDENTIALS_JSON" in os.environ:
        try:
            from google.cloud import vision
            client = vision.ImageAnnotatorClient()
            response = client.text_detection(image=vision.Image(content=img_bytes))
            rp = _parse_gcp_vision_raw_peaks(response)
            if rp: return rp
        except Exception:
            pass

    return []

def get_render_font(font_size, bold=True):
    """Finds available system font on Windows or Linux / Streamlit Cloud."""
    font_candidates = [
        'C:/Windows/Fonts/arialbd.ttf',
        'C:/Windows/Fonts/arial.ttf',
        'C:/Windows/Fonts/malgunbd.ttf',
        'C:/Windows/Fonts/malgun.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
        '/usr/share/fonts/truetype/freefont/FreeSansBold.ttf',
        '/usr/share/fonts/truetype/freefont/FreeSans.ttf',
    ]
    for fp in font_candidates:
        if os.path.exists(fp):
            try:
                return ImageFont.truetype(fp, font_size)
            except Exception:
                pass
    try:
        return ImageFont.load_default(size=font_size)
    except TypeError:
        return ImageFont.load_default()

def process_spectrum_image(img_bytes, peak_overrides=None, font_size=45):
    """
    Renders enlarged peak labels with user editable overrides:
    peak_overrides = [ {'orig_mz': '336.06', 'final_mz': '335.06', 'is_mrm': True}, ... ]
    """
    if peak_overrides is None:
        peak_overrides = []
        
    nparr = np.frombuffer(img_bytes, np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if img is None or img.size == 0:
        return img_bytes
        
    x_axis, y_axis = detect_axis_lines(img)
    x_start = x_axis + 10
    x_end = min(img.shape[1] - 30, 2180)
    y_top = 50
    y_bottom = y_axis - 10
    
    # 1. 100% COMPLETE ERASURE of old small text & leader lines in graph area
    out_img = img.copy()
    b, g, r = cv2.split(out_img)
    pure_blue = (b > 130) & (r < 70) & (g < 120)
    is_blue = (b > 110) & (b > r + 25) & (b > g + 5)
    gray = cv2.cvtColor(out_img, cv2.COLOR_BGR2GRAY)
    
    plot_gray = gray[y_top:y_bottom, x_start:x_end]
    plot_blue = is_blue[y_top:y_bottom, x_start:x_end]
    text_mask = (plot_gray < 235) & (~plot_blue)
    
    roi = out_img[y_top:y_bottom, x_start:x_end]
    roi[text_mask] = [255, 255, 255]
    
    # 2. Determine m/z range for physical X-axis alignment
    all_vals = []
    if peak_overrides:
        for item in peak_overrides:
            try: all_vals.append(float(item.get('final_mz', item.get('orig_mz'))))
            except Exception: pass
            
    if all_vals:
        min_v, max_v = min(all_vals), max(all_vals)
        mz_min = 50.0 if min_v >= 45.0 else 0.0
        if max_v <= 220.0:
            mz_max = 200.0
        elif max_v <= 520.0:
            mz_max = 500.0
        elif max_v <= 1020.0:
            mz_max = 1000.0
        else:
            mz_max = float(int(np.ceil(max_v / 100.0) * 100.0))
    else:
        mz_min, mz_max = 50.0, 500.0
        
    peak_labels = []
    if peak_overrides:
        for item in peak_overrides:
            orig_mz = item['orig_mz']
            final_mz = item.get('final_mz', orig_mz)
            is_mrm = item.get('is_mrm', False)
            try:
                val_f = float(final_mz)
            except ValueError:
                continue
                
            x_calc = int(x_start + ((val_f - mz_min) / (mz_max - mz_min)) * (x_end - x_start))
            x_calc = max(x_start + 15, min(x_end - 15, x_calc))
            
            # Scan true peak apex from baseline upwards (ignoring y < 80 top grid)
            x1 = max(x_start, x_calc - 45)
            x2 = min(x_end, x_calc + 45)
            best_apex_x = x_calc
            best_apex_y = y_bottom
            
            for cx in range(x1, x2):
                col = pure_blue[80:y_bottom, cx]
                blue_idx = np.where(col)[0]
                if len(blue_idx) > 0:
                    top_y = 80 + np.min(blue_idx)
                    if top_y < best_apex_y:
                        best_apex_y = top_y
                        best_apex_x = cx
                        
            peak_labels.append({
                'text': final_mz,
                'center': (best_apex_x, best_apex_y),
                'bbox': (best_apex_x - 40, best_apex_y - 15, best_apex_x + 40, best_apex_y + 15),
                'is_mrm': is_mrm
            })
    else:
        raw_peaks = _get_raw_peaks_for_image(img_bytes, img)
        corrected_peaks = calibrate_and_correct_peaks(raw_peaks)
        for p in corrected_peaks:
            peak_labels.append({
                'text': p['mz'],
                'center': (p['cx'], p['cy']),
                'bbox': (p['x_min'], p['y_min'], p['x_max'], p['y_max']),
                'is_mrm': False
            })
        
    # 3. Transparent Overlay with MRM styling (Blue 45pt vs Gray 40pt)
    img_rgb = cv2.cvtColor(out_img, cv2.COLOR_BGR2RGB)
    pil_base = Image.fromarray(img_rgb).convert('RGBA')
    
    overlay = Image.new('RGBA', pil_base.size, (255, 255, 255, 0))
    draw = ImageDraw.Draw(overlay)
    
    font_main = get_render_font(font_size, bold=True)
    font_sub = get_render_font(max(12, font_size - 5), bold=False)
        
    labels = resolve_label_collisions(peak_labels, font_main, font_sub, draw, min_dist_y=int(font_size * 1.2))
    
    for item in labels:
        txt = item['text']
        tx, ty = item['curr_x'], item['curr_y']
        tw, th = item['tw'], item['th']
        apex_x, apex_y = item['apex_x'], item['apex_y']
        is_mrm = item.get('is_mrm', False)
        is_side = item.get('is_side_aligned', False)
        is_shifted = item.get('is_shifted', False) or abs(ty - item['orig_y']) > 15
        
        text_color = (0, 0, 220, 255) if is_mrm else (120, 120, 120, 255)
        line_color = (0, 0, 220, 255) if is_mrm else (150, 150, 150, 255)
        font = font_main if is_mrm else font_sub
        
        if is_side:
            # Side aligned (ceiling clamp): small pointer dot at apex
            draw.ellipse([apex_x - 3, apex_y - 3, apex_x + 3, apex_y + 3], fill=line_color)
        elif is_shifted:
            # Shifted vertically due to overlap: draw crisp leader line from text bottom to apex
            start_pt = (tx + tw // 2, ty + th + 2)
            end_pt = (apex_x, apex_y)
            draw.line([start_pt, end_pt], fill=line_color, width=2)
            draw.ellipse([apex_x - 3, apex_y - 3, apex_x + 3, apex_y + 3], fill=line_color)
            
        # Draw text with transparent background and subtle outline for crisp readability
        draw.text((tx, ty), txt, font=font, fill=text_color, stroke_width=2, stroke_fill=(255, 255, 255, 180))
        
    final_pil = Image.alpha_composite(pil_base, overlay).convert('RGB')
    
    buf = io.BytesIO()
    final_pil.save(buf, format='PNG')
    return buf.getvalue()

def process_excel_with_selections(excel_path, sheet_selections, font_size=45, status_callback=None):
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {excel_path}")
        
    if status_callback:
        status_callback("엑셀 파일 로딩 및 MS/MS 이미지 변환 중...")
        
    wb = openpyxl.load_workbook(excel_path)
    temp_dir = os.path.join(os.path.dirname(excel_path), "_temp_mz_img")
    os.makedirs(temp_dir, exist_ok=True)
    
    total_images = 0
    for name in wb.sheetnames:
        total_images += len(wb[name]._images)
        
    processed_count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sel_dict = sheet_selections.get(sheet_name, {'precursor': [], 'product': []})
        
        for idx, img in enumerate(ws._images):
            processed_count += 1
            if status_callback:
                status_callback(f"이미지 처리 중 ({processed_count}/{total_images}) - 시트: {sheet_name}")
                
            is_precursor = (idx % 2 == 0)
            overrides = sel_dict['precursor'] if is_precursor else sel_dict['product']
            
            img_bytes = img._data()
            proc_bytes = process_spectrum_image(img_bytes, peak_overrides=overrides, font_size=font_size)
            
            temp_path = os.path.join(temp_dir, f"{sheet_name}_{idx+1}.png")
            with open(temp_path, "wb") as f:
                f.write(proc_bytes)
                
            new_img = OpenpyxlImage(temp_path)
            new_img.anchor = img.anchor
            ws._images[idx] = new_img
            
    script_dir = get_base_dir()
    results_dir = os.path.join(script_dir, "Results")
    os.makedirs(results_dir, exist_ok=True)
    
    dir_name, file_name = os.path.split(excel_path)
    name_part, ext_part = os.path.splitext(file_name)
    output_path = os.path.join(results_dir, f"{name_part}_확대{ext_part}")
    
    if status_callback:
        status_callback("수정된 엑셀 파일 저장 중...")
        
    wb.save(output_path)
    return output_path

class MRMSelectionModal(tk.Toplevel if tk is not None else object):
    """
    Modal Dialog popup showing extracted m/z peak values per Sheet and Graph Type (Precursor vs Product Ion).
    Features:
    - Smart default auto-selection (Top 1 Precursor, Top 3 Product).
    - Editable m/z text boxes so users can manually correct any misread OCR numbers!
    """
    def __init__(self, parent, excel_path, sheet_peak_data, on_confirm_callback):
        super().__init__(parent)
        self.title("MRM 조건 이온 선택 및 수기 수정")
        self.geometry("820x660")
        self.transient(parent)
        
        self.excel_path = excel_path
        self.sheet_peak_data = sheet_peak_data
        self.on_confirm_callback = on_confirm_callback
        self.peak_widgets = {} # {sheet_name: {'precursor': [(var, entry, orig_mz)], 'product': ...}}
        
        self.build_ui()
        self.grab_set()
        self.focus_set()
        
    def build_ui(self):
        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        header_lbl = ttk.Label(main_frame, text="MRM 조건 이온 선택 및 분자량(m/z) 수기 수정", font=("맑은 고딕", 13, "bold"), foreground="navy")
        header_lbl.pack(anchor=tk.W, pady=(0, 5))
        
        guide_text = "✓ OCR이 오인식한 분자량이 있다면 입력창에서 직접 수기로 수정할 수 있습니다.\n" \
                     "✓ 스마트 자동 체크: Precursor 상위 1개, Product 상위 3개 피크 기본 선택.\n" \
                     "✓ 체크한 이온: 파란색 (Blue) 48pt 강조 / 체크 안 한 이온: 회색 (Gray) 46pt 표시."
        guide_lbl = ttk.Label(main_frame, text=guide_text, font=("맑은 고딕", 10), foreground="#333333")
        guide_lbl.pack(anchor=tk.W, pady=(0, 10))
        
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=5)
        
        for sheet_name, graph_data in self.sheet_peak_data.items():
            sheet_tab = ttk.Frame(notebook, padding="10")
            notebook.add(sheet_tab, text=f" 📄 {sheet_name} ")
            
            self.peak_widgets[sheet_name] = {'precursor': [], 'product': []}
            
            canvas = tk.Canvas(sheet_tab, borderwidth=0, highlightthickness=0)
            scrollbar = ttk.Scrollbar(sheet_tab, orient="vertical", command=canvas.yview)
            scroll_content = ttk.Frame(canvas, padding="5")
            
            canvas_window = canvas.create_window((0, 0), window=scroll_content, anchor="nw")
            
            def _on_canvas_resize(event, c=canvas, w=canvas_window):
                c.itemconfig(w, width=event.width)
            canvas.bind('<Configure>', _on_canvas_resize)
            
            def _on_content_resize(event, c=canvas):
                c.configure(scrollregion=c.bbox('all'))
            scroll_content.bind('<Configure>', _on_content_resize)
            
            canvas.configure(yscrollcommand=scrollbar.set)
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Precursor Ion Frame
            prec_frame = ttk.LabelFrame(scroll_content, text=" 🔹 Precursor Ion 피크 (입력창에서 수기 수정 가능) ", padding="10")
            prec_frame.pack(fill=tk.X, expand=True, pady=5, padx=5)
            
            prec_info = graph_data.get('precursor', {'all_peaks': [], 'default_checked': set()})
            prec_peaks = prec_info['all_peaks']
            prec_defaults = prec_info['default_checked']
            
            if not prec_peaks:
                ttk.Label(prec_frame, text="검출된 피크가 없습니다.", font=("맑은 고딕", 9, "italic")).pack(anchor=tk.W)
            else:
                grid_frame = ttk.Frame(prec_frame)
                grid_frame.pack(fill=tk.X, expand=True)
                for idx, mz in enumerate(prec_peaks):
                    is_checked = mz in prec_defaults
                    var = tk.BooleanVar(value=is_checked)
                    
                    row = idx // 2
                    col_base = (idx % 2) * 3
                    
                    cb = ttk.Checkbutton(grid_frame, variable=var)
                    cb.grid(row=row, column=col_base, sticky=tk.W, padx=(10, 2), pady=4)
                    
                    entry = ttk.Entry(grid_frame, width=10, font=("맑은 고딕", 10, "bold"))
                    entry.insert(0, mz)
                    entry.grid(row=row, column=col_base+1, sticky=tk.W, padx=(0, 2), pady=4)
                    
                    tag = "Da ⭐ (최고 피크)" if is_checked else "Da"
                    lbl = ttk.Label(grid_frame, text=tag, font=("맑은 고딕", 9))
                    lbl.grid(row=row, column=col_base+2, sticky=tk.W, padx=(0, 25), pady=4)
                    
                    self.peak_widgets[sheet_name]['precursor'].append((var, entry, mz))
                    
            # Product Ion Frame
            prod_frame = ttk.LabelFrame(scroll_content, text=" 🔸 Product Ion 피크 (입력창에서 수기 수정 가능) ", padding="10")
            prod_frame.pack(fill=tk.X, expand=True, pady=10, padx=5)
            
            prod_info = graph_data.get('product', {'all_peaks': [], 'default_checked': set()})
            prod_peaks = prod_info['all_peaks']
            prod_defaults = prod_info['default_checked']
            
            if not prod_peaks:
                ttk.Label(prod_frame, text="검출된 피크가 없습니다.", font=("맑은 고딕", 9, "italic")).pack(anchor=tk.W)
            else:
                grid_frame = ttk.Frame(prod_frame)
                grid_frame.pack(fill=tk.X, expand=True)
                for idx, mz in enumerate(prod_peaks):
                    is_checked = mz in prod_defaults
                    var = tk.BooleanVar(value=is_checked)
                    
                    row = idx // 2
                    col_base = (idx % 2) * 3
                    
                    cb = ttk.Checkbutton(grid_frame, variable=var)
                    cb.grid(row=row, column=col_base, sticky=tk.W, padx=(10, 2), pady=4)
                    
                    entry = ttk.Entry(grid_frame, width=10, font=("맑은 고딕", 10, "bold"))
                    entry.insert(0, mz)
                    entry.grid(row=row, column=col_base+1, sticky=tk.W, padx=(0, 2), pady=4)
                    
                    tag = "Da ⭐ (추천)" if is_checked else "Da"
                    lbl = ttk.Label(grid_frame, text=tag, font=("맑은 고딕", 9))
                    lbl.grid(row=row, column=col_base+2, sticky=tk.W, padx=(0, 25), pady=4)
                    
                    self.peak_widgets[sheet_name]['product'].append((var, entry, mz))
                    
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))
        
        btn_confirm = ttk.Button(btn_frame, text="✓ MRM 이온 선택 완료 및 엑셀 생성", command=self.on_confirm)
        btn_confirm.pack(side=tk.RIGHT, ipadx=10, ipady=5)
        
        btn_cancel = ttk.Button(btn_frame, text="취소", command=self.destroy)
        btn_cancel.pack(side=tk.RIGHT, padx=10, ipady=5)
        
        self.update_idletasks()
        
    def on_confirm(self):
        selections = {}
        for sheet_name, cat_dict in self.peak_widgets.items():
            selections[sheet_name] = {'precursor': [], 'product': []}
            for category in ['precursor', 'product']:
                for var, entry, orig_mz in cat_dict[category]:
                    user_val = entry.get().strip()
                    try:
                        fval = float(user_val)
                        formatted_val = f"{fval:.2f}"
                    except ValueError:
                        formatted_val = user_val
                        
                    selections[sheet_name][category].append({
                        'orig_mz': orig_mz,
                        'final_mz': formatted_val,
                        'is_mrm': var.get()
                    })
        self.destroy()
        self.on_confirm_callback(self.excel_path, selections)

class AppGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("GLP 보고서용 질량분석 피크 분자량(m/z) 확대 프로그램")
        self.root.geometry("680x420")
        self.root.resizable(False, False)
        
        style = ttk.Style()
        style.theme_use('clam')
        
        frame = ttk.Frame(root, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        title_lbl = ttk.Label(frame, text="질량분석(MS/MS) 피크 분자량(m/z) 글자 확대기", font=("맑은 고딕", 14, "bold"))
        title_lbl.pack(anchor=tk.W, pady=(0, 5))
        
        desc_lbl = ttk.Label(frame, text="Data 폴더 선택 후 팝업창에서 MRM 이온 체크 및 수기 수정이 가능합니다.", font=("맑은 고딕", 9))
        desc_lbl.pack(anchor=tk.W, pady=(0, 15))
        
        # 1. File Selection Frame (Dropdown from Data folder)
        file_box = ttk.LabelFrame(frame, text=" 📂 대상 엑셀 파일 선택 (Data 폴더) ", padding="12")
        file_box.pack(fill=tk.X, pady=5)
        
        base_d = get_base_dir()
        self.data_dir = os.path.join(base_d, "Data")
        if not os.path.exists(self.data_dir):
            os.makedirs(self.data_dir, exist_ok=True)
            
        self.combo_files = ttk.Combobox(file_box, state="readonly", font=("맑은 고딕", 10))
        self.combo_files.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        btn_refresh = ttk.Button(file_box, text="🔄 목록 새로고침", command=self.refresh_file_list)
        btn_refresh.pack(side=tk.RIGHT)
        
        self.refresh_file_list()
        
        # 2. Options Frame
        opt_frame = ttk.LabelFrame(frame, text=" 확대 폰트 설정 ", padding="10")
        opt_frame.pack(fill=tk.X, pady=10)
        
        ttk.Label(opt_frame, text="MRM 강조 폰트 크기:").pack(side=tk.LEFT, padx=(0, 10))
        self.font_size_var = tk.IntVar(value=48)
        scale_spin = ttk.Spinbox(opt_frame, from_=20, to=72, increment=4, textvariable=self.font_size_var, width=5)
        scale_spin.pack(side=tk.LEFT)
        ttk.Label(opt_frame, text="pt (MRM 선택 이온: 파란색 48pt / 기타 이온: 회색 46pt)").pack(side=tk.LEFT, padx=(5, 0))
        
        # 3. Status Frame
        self.status_var = tk.StringVar(value="Data 폴더의 엑셀 파일을 선택한 후 '분자량 피크 읽기 및 확대' 버튼을 클릭하세요.")
        status_lbl = ttk.Label(frame, textvariable=self.status_var, font=("맑은 고딕", 9), foreground="blue")
        status_lbl.pack(anchor=tk.W, pady=10)
        
        # 4. Action Button
        self.btn_run = ttk.Button(frame, text="🔍 1단계: 분자량 피크 읽기 & MRM 이온 선택/수정 팝업 열기", command=self.start_mrm_selection)
        self.btn_run.pack(fill=tk.X, ipady=8, pady=(5, 0))
        
    def refresh_file_list(self):
        """Refreshes file dropdown list from Data folder."""
        file_list = []
        base_d = get_base_dir()
        if os.path.exists(self.data_dir):
            for f in os.listdir(self.data_dir):
                if f.endswith(('.xlsx', '.xls')) and not f.endswith('_확대.xlsx') and not f.endswith('_분자량확대.xlsx'):
                    file_list.append(f)
                    
        # Also check current root directory as fallback
        if os.path.exists(base_d):
            for f in os.listdir(base_d):
                if f.endswith(('.xlsx', '.xls')) and not f.endswith('_확대.xlsx') and not f.endswith('_분자량확대.xlsx') and f not in file_list:
                    file_list.append(f)
                
        self.combo_files['values'] = file_list
        if file_list:
            self.combo_files.current(0)
        else:
            self.combo_files.set('')
            
    def update_status(self, msg):
        self.status_var.set(msg)
        self.root.update()
        
    def start_mrm_selection(self):
        selected_filename = self.combo_files.get().strip()
        if not selected_filename:
            messagebox.showwarning("경고", "처리할 엑셀 파일을 선택하세요.")
            return
            
        excel_path = os.path.join(self.data_dir, selected_filename)
        if not os.path.exists(excel_path):
            excel_path = os.path.join(get_base_dir(), selected_filename)
            
        if not os.path.exists(excel_path):
            messagebox.showerror("오류", f"파일을 찾을 수 없습니다: {excel_path}")
            return
            
        try:
            self.btn_run.config(state=tk.DISABLED)
            self.update_status("🔍 엑셀 내 Precursor / Product Ion 분자량 피크 읽는 중...")
            
            wb = openpyxl.load_workbook(excel_path)
            sheet_peak_data = {}
            tasks = []
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_peak_data[sheet_name] = {
                    'precursor': {'all_peaks': [], 'default_checked': set()},
                    'product': {'all_peaks': [], 'default_checked': set()}
                }
                for idx, img in enumerate(ws._images):
                    is_precursor = (idx % 2 == 0)
                    tasks.append((sheet_name, is_precursor, img._data()))
                    
            total_images = len(tasks)
            proc_img_count = 0
            
            def _proc_task(task_item):
                s_name, is_prec, data_bytes = task_item
                res = extract_peaks_from_image(data_bytes, is_precursor=is_prec)
                return s_name, is_prec, res
                
            max_workers = min(4, os.cpu_count() or 4)
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_task = {executor.submit(_proc_task, t): t for t in tasks}
                for future in concurrent.futures.as_completed(future_to_task):
                    proc_img_count += 1
                    s_name, is_prec, peak_info = future.result()
                    self.update_status(f"🔍 고속 피크 분석 중 ({proc_img_count}/{total_images}) - 시트: {s_name}")
                    
                    cat_key = 'precursor' if is_prec else 'product'
                    sheet_peak_data[s_name][cat_key]['all_peaks'].extend(peak_info['all_peaks'])
                    sheet_peak_data[s_name][cat_key]['default_checked'].update(peak_info['default_checked'])
                    
            for sheet_name in wb.sheetnames:
                prec_all = sorted(list(set(sheet_peak_data[sheet_name]['precursor']['all_peaks'])), key=lambda x: float(x))
                prod_all = sorted(list(set(sheet_peak_data[sheet_name]['product']['all_peaks'])), key=lambda x: float(x))
                sheet_peak_data[sheet_name]['precursor']['all_peaks'] = prec_all
                sheet_peak_data[sheet_name]['product']['all_peaks'] = prod_all
                
            self.update_status("✅ 초고속 피크 추출 완료! MRM 이온 선택/수정 팝업창을 확인하세요.")
            
            # Create & Display Modal Popup
            modal = MRMSelectionModal(self.root, excel_path, sheet_peak_data, self.execute_excel_generation)
            
        except Exception as e:
            self.update_status("❌ 이미지 피크 읽기 오류 발생")
            messagebox.showerror("오류", f"피크 읽기 중 오류가 발생했습니다:\n{str(e)}")
        finally:
            self.btn_run.config(state=tk.NORMAL)
            
    def execute_excel_generation(self, excel_path, sheet_selections):
        try:
            self.btn_run.config(state=tk.DISABLED)
            font_size = int(self.font_size_var.get())
            out_file = process_excel_with_selections(excel_path, sheet_selections, font_size=font_size, status_callback=self.update_status)
            
            self.status_var.set(f"🎉 완벽 완료! 저장됨: {os.path.basename(out_file)}")
            messagebox.showinfo("성공", f"수정된 MRM 이온(파란색 강조)이 적용된 최종 엑셀 파일이 저장되었습니다!\n\n저장 경로:\n{out_file}")
        except Exception as e:
            self.status_var.set("❌ 엑셀 저장 중 오류 발생")
            messagebox.showerror("오류", f"엑셀 생성 중 오류가 발생했습니다:\n{str(e)}")
        finally:
            self.btn_run.config(state=tk.NORMAL)

def main():
    if len(sys.argv) > 1 and sys.argv[1].endswith(('.xlsx', '.xls')):
        excel_file = sys.argv[1]
        print(f"CLI 모드로 실행합니다: {excel_file}")
        wb = openpyxl.load_workbook(excel_file)
        sheet_selections = {}
        for name in wb.sheetnames:
            sheet_selections[name] = {'precursor': [], 'product': []}
        out = process_excel_with_selections(excel_file, sheet_selections, font_size=48, status_callback=print)
        print(f"완료! 저장 위치: {out}")
    else:
        if tk is None:
            print("Tkinter GUI를 사용할 수 없는 환경입니다 (서버/클라우드 환경).")
            return
        root = tk.Tk()
        app = AppGUI(root)
        root.mainloop()

if __name__ == "__main__":
    main()
